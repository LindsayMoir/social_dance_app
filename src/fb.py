"""
fb.py

This module defines the FacebookEventScraper class for scraping Facebook event data
using Playwright and BeautifulSoup. It handles logging into Facebook, extracting event
link and text, processing URLs, and 
interacting with a database and Language Learning Model (LLM) to process and store
event data.

Classes:
    FacebookEventScraper:
        - Initializes with configuration, sets up Playwright for browser automation.
        - Logs into Facebook and maintains a session for scraping.
        - Extracts event link and content from Facebook event pages.
        - Processes Facebook URLs including fixing malformed URLs.
        - Interacts with LLMHandler for natural language processing tasks.
        - Interacts with DatabaseHandler to read/write URLs and events to the database.
        - Provides multiple driver methods for different scraping workflows:
            • driver_fb_urls: Processes Facebook URLs from the database.
            • driver_fb_search: Searches for Facebook events based on keywords.
            • driver_no_urls: Handles events without URLs, attempts to find and update them.

Usage Example:
    if __name__ == "__main__":
        # Load configuration and configure logging
        from logging_config import setup_logging
        setup_logging('fb')
        with open('config/config.yaml', 'r') as file:
            config = yaml.safe_load(file)

        # Initialize dependencies
        llm_handler = LLMHandler(config_path='config/config.yaml')
        db_handler = llm_handler.db_handler  # Use the DatabaseHandler from LLMHandler
        fb_scraper = FacebookEventScraper(config_path='config/config.yaml')

        # Run scraping drivers
        fb_scraper.driver_fb_urls()
        fb_scraper.driver_fb_search()
        fb_scraper.driver_no_urls()

        # Clean up resource
        fb_scraper.browser.close()
        fb_scraper.playwright.stop()
        
        # Logging of process duration handled in __main__ block

Dependencies:
    - Playwright: For browser automation to navigate and scrape Facebook.
    - BeautifulSoup: To parse HTML content of event pages.
    - pandas: For data manipulation and CSV operations.
    - fuzzywuzzy: For fuzzy string matching.
    - requests: For HTTP requests when scraping non-JS rendered pages.
    - yaml: For configuration file parsing.
    - logging: For tracking the execution flow and errors.
    - re: For regular expression operations.
    - Other custom modules: llm (LLMHandler), db (DatabaseHandler).

Note:
    - The module assumes valid configuration in 'config/config.yaml'.
    - Logging is configured in the main section to record key actions and errors.
    - The class methods heavily rely on external services (Facebook, database, LLM),
      and their correct functioning depends on valid credentials and network access.
"""


from bs4 import BeautifulSoup
from datetime import datetime
from rapidfuzz import fuzz
import logging
from openpyxl import load_workbook
import os
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import random
import re
from sqlalchemy import text
import subprocess
import time
from urllib.parse import urlparse, parse_qs, unquote
import yaml

# Import other classes
from credentials import get_credentials
from db import DatabaseHandler
from llm import LLMHandler
from logging_utils import log_extracted_text
from secret_paths import get_auth_file

# Get config
with open('config/config.yaml', 'r') as file:
    config = yaml.safe_load(file)

# Global handlers for class methods (declared at module level, initialized in main())
llm_handler = None
db_handler = None


def should_use_fb_checkpoint(config_data: dict, is_render: bool) -> bool:
    """
    Decide whether fb.py should read facebook URLs from checkpoint CSV.
    """
    if is_render:
        return False
    return bool(config_data.get('checkpoint', {}).get('fb_urls_cp_status', False))


def canonicalize_facebook_url(url: str) -> str:
    """
    Normalize Facebook login redirect URLs to their real target page.
    """
    if not url:
        return url

    current = str(url).strip()
    for _ in range(3):  # safety guard for nested next= redirects
        parsed = urlparse(current)
        host = (parsed.netloc or "").lower()
        path = (parsed.path or "").lower()

        if "facebook.com" not in host:
            return current

        qs = parse_qs(parsed.query)
        should_unwrap_next = "/login" in path or "/recover/initiate" in path
        if not should_unwrap_next:
            return current

        nxt = qs.get("next", [None])[0]
        if not nxt:
            return current

        decoded = unquote(nxt).strip()
        if not decoded:
            return current
        if decoded.startswith("/"):
            decoded = f"https://www.facebook.com{decoded}"
        if decoded == current:
            return current
        current = decoded
    return current


def is_facebook_login_redirect(url: str) -> bool:
    """
    Return True if URL points to a Facebook login endpoint.
    """
    parsed = urlparse(url or "")
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()
    return "facebook.com" in host and "/login" in path


def is_non_content_facebook_url(url: str) -> bool:
    """
    Return True for Facebook utility/auth/share endpoints that are not event/group content pages.
    """
    parsed = urlparse(url or "")
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()

    if "facebook.com" not in host:
        return False

    blocked_prefixes = (
        "/sharer/",
        "/dialog/",
        "/recover/",
        "/share.php",
    )
    if path.startswith(blocked_prefixes):
        return True
    if "sharer.php" in path:
        return True
    return False


def classify_facebook_access_state(current_url: str, page_content: str) -> str:
    """
    Classify Facebook page access state as one of:
    - 'ok': content appears accessible
    - 'login': authentication wall detected
    - 'blocked': temporary block/rate-limit/interstitial detected
    """
    url_lower = (current_url or "").lower()
    content_lower = (page_content or "").lower()

    # Treat explicit auth URLs as login walls.
    if any(m in url_lower for m in ("/login", "/checkpoint/")):
        return "login"

    # Use stricter content checks to avoid false positives on pages that contain generic "log in" text.
    has_login_form = ('name="email"' in content_lower and 'name="pass"' in content_lower)
    has_login_gate = ("must log in" in content_lower or "two-factor" in content_lower)
    if has_login_form or has_login_gate:
        return "login"

    blocked_markers = (
        "temporarily blocked",
        "misusing this feature",
        "too many requests",
        "try again later",
        "rate limited",
    )
    if "/sorry/" in url_lower or any(m in content_lower for m in blocked_markers):
        return "blocked"

    return "ok"


def sanitize_facebook_seed_urls(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Normalize and filter Facebook seed URLs before processing.

    This guards against stale checkpoint rows containing login/share/recover/dialog
    endpoints and de-duplicates rows after canonicalization.
    """
    stats = {
        "input_rows": len(df),
        "empty_rows_dropped": 0,
        "canonicalized_rows": 0,
        "non_content_rows_dropped": 0,
        "duplicate_rows_dropped": 0,
        "output_rows": 0,
    }

    if "link" not in df.columns:
        stats["output_rows"] = len(df)
        return df, stats

    cleaned = df.copy()
    cleaned["link"] = cleaned["link"].astype(str).str.strip()
    before_empty = len(cleaned)
    cleaned = cleaned[cleaned["link"] != ""].copy()
    stats["empty_rows_dropped"] = before_empty - len(cleaned)

    cleaned["normalized_link"] = cleaned["link"].apply(canonicalize_facebook_url).astype(str).str.strip()
    stats["canonicalized_rows"] = int((cleaned["normalized_link"] != cleaned["link"]).sum())

    is_non_content = cleaned["normalized_link"].apply(is_non_content_facebook_url)
    stats["non_content_rows_dropped"] = int(is_non_content.sum())
    cleaned = cleaned[~is_non_content].copy()

    cleaned["link"] = cleaned["normalized_link"]
    cleaned = cleaned.drop(columns=["normalized_link"])

    before_dedup = len(cleaned)
    cleaned = cleaned.drop_duplicates(subset=["link"], keep="first").reset_index(drop=True)
    stats["duplicate_rows_dropped"] = before_dedup - len(cleaned)
    stats["output_rows"] = len(cleaned)
    return cleaned, stats


def get_git_revision() -> str:
    """
    Return current git short SHA for runtime traceability.
    """
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
    except Exception:
        return "unknown"


class FacebookEventScraper():
    def __init__(self, config_path: str = "config/config.yaml") -> None:
        # Load configuration
        with open(config_path, 'r') as file:
            self.config = yaml.safe_load(file)

        # Start Playwright
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=self.config['crawling']['headless'])

        # Create a single context & page, reusing 'facebook_auth.json'
        # Uses Render Secret Files path if available, otherwise local path
        self.facebook_auth_path = get_auth_file('facebook')
        self.context = self.browser.new_context(storage_state=self.facebook_auth_path)
        self.page = self.context.new_page()
        # keep a stable reference for re‑use
        self.logged_in_page = self.page

        # Attempt login
        self.login_success = self.login_to_facebook()
        if self.login_success:
            logging.info("Facebook login successful.")
        else:
            logging.error("Facebook login failed. Aborting run to avoid zero-yield scraping.")
            self._safe_shutdown_browser()
            raise RuntimeError("Facebook login failed; stopping fb.py.")

        # Run statistics tracking
        if config['testing']['status']:
            self.run_name = f"Test Run {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            self.run_description = "Test Run Description"
        else:
            self.run_name = "Facebook Event Scraper Run"
            self.run_description = f"Production {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # **Tracking Variables**
        self.unique_urls = set()  # Store unique URLs
        self.unique_urls_count = 0  # Unique URLs processed
        self.total_url_attempts = 0  # Total number of URL contact attempts
        self.urls_with_extracted_text = 0  # URLs where text was successfully extracted
        self.urls_with_found_keywords = 0  # URLs where keywords were found
        self.events_written_to_db = 0  # Number of events successfully written to DB

        # Initialize run statistics
        self.start_time = datetime.now()
        self.total_url_attempts = 0
        self.urls_with_extracted_text = 0
        self.urls_with_found_keywords = 0
        self.events_written_to_db = 0

        # URL tracking
        self.urls_visited = set()

        # Get keywords (deferred if handlers not initialized)
        self.keywords_list = llm_handler.get_keywords() if llm_handler else []


    def _safe_shutdown_browser(self) -> None:
        """Close Playwright resources safely."""
        try:
            if hasattr(self, 'context') and self.context:
                self.context.close()
        except Exception:
            pass
        try:
            if hasattr(self, 'browser') and self.browser:
                self.browser.close()
        except Exception:
            pass
        try:
            if hasattr(self, 'playwright') and self.playwright:
                self.playwright.stop()
        except Exception:
            pass


    def _ensure_authenticated_or_raise(self) -> None:
        """Fail fast if Facebook authentication was not established."""
        if not getattr(self, 'login_success', False):
            raise RuntimeError("Facebook session is not authenticated; aborting run.")


    def _retry_headless_login_with_fresh_context(self) -> bool:
        """
        Repair attempt: rebuild context without storage state and retry login once.
        """
        try:
            if hasattr(self, 'context') and self.context:
                self.context.close()
        except Exception:
            pass

        try:
            self.context = self.browser.new_context()
            self.page = self.context.new_page()
            self.logged_in_page = self.page
            page = self.logged_in_page

            page.goto("https://www.facebook.com/login", wait_until="domcontentloaded", timeout=20000)
            page.wait_for_selector("input[name='email']", timeout=10000)
            page.wait_for_selector("input[name='pass']", timeout=10000)

            email, password, _ = get_credentials("Facebook")
            page.fill("input[name='email']", email)
            page.fill("input[name='pass']", password)
            page.click("button[type='submit']")
            page.wait_for_timeout(5000)

            if "login" in page.url.lower():
                return False

            self.context.storage_state(path=self.facebook_auth_path)
            from secret_paths import sync_auth_to_db
            sync_auth_to_db(self.facebook_auth_path, 'facebook')
            return True
        except Exception as e:
            logging.error(f"_retry_headless_login_with_fresh_context: failed: {e}")
            return False
    

    def login_to_facebook(self) -> bool:
        """
        Ensure we're logged into Facebook using the existing page.
        If already logged in, return immediately; otherwise:
          - In headless=False, prompt for manual login (credentials + 2FA), then CAPTCHA.
          - In headless=True, submit credentials programmatically.
        Returns:
            bool: True if login was successful, False otherwise.
        """
        page = self.logged_in_page
        headless = self.config['crawling'].get('headless', True)

        # 1) Navigate to the login page
        try:
            page.goto(
                "https://www.facebook.com/login",
                wait_until="networkidle",
                timeout=random.randint(20000 // 2, int(20000 * 1.5))
            )
        except PlaywrightTimeoutError:
            logging.warning("login_to_facebook: login page load timed out; proceeding.")

        # 2) If already authenticated, done
        if "login" not in page.url.lower():
            logging.info("login_to_facebook: already authenticated.")
            return True

        # 3) Manual flow (visible browser)
        if not headless:
            print("\n=== MANUAL FACEBOOK LOGIN ===")
            print("1) In the browser window, enter your username/password and complete any 2FA.")
            input("   Once you've logged in successfully, press ENTER here to continue… ")
            try:
                page.reload(wait_until="networkidle", timeout=20000)
            except PlaywrightTimeoutError:
                logging.warning("login_to_facebook: reload after manual login timed out; continuing.")

            # CAPTCHA detection using centralized handler
            from captcha_handler import CaptchaHandler
            captcha_detected = CaptchaHandler.detect_and_handle_sync(page, "Facebook", timeout=5000)
            if captcha_detected:
                try:
                    page.reload(wait_until="networkidle", timeout=20000)
                except PlaywrightTimeoutError:
                    logging.warning("login_to_facebook: reload after CAPTCHA timed out; continuing.")

            if "login" in page.url.lower():
                logging.error("login_to_facebook: still on login page after manual flow.")
                return False

            # Persist state
            try:
                self.context.storage_state(path=self.facebook_auth_path)
                logging.info("login_to_facebook: session state saved (manual).")
                # Sync to database
                from secret_paths import sync_auth_to_db
                sync_auth_to_db(self.facebook_auth_path, 'facebook')
            except Exception as e:
                logging.warning(f"login_to_facebook: could not save session state: {e}")
            return True

        # 4) Automated flow (headless)
        try:
            page.wait_for_selector("input[name='email']", timeout=10000)
            page.wait_for_selector("input[name='pass']", timeout=10000)
        except PlaywrightTimeoutError:
            # Sometimes Facebook redirects to an authenticated state or an interstitial.
            try:
                page.goto("https://www.facebook.com/events", wait_until="domcontentloaded", timeout=15000)
                if "login" not in page.url.lower():
                    logging.info("login_to_facebook: authenticated session detected after selector timeout.")
                    return True
            except Exception:
                pass

            logging.warning("login_to_facebook: login form did not appear. Retrying with fresh context.")
            if self._retry_headless_login_with_fresh_context():
                logging.info("login_to_facebook: recovered session with fresh context.")
                return True

            logging.error("login_to_facebook: login form did not appear and retry failed.")
            return False

        try:
            email, password, _ = get_credentials("Facebook")
            page.fill("input[name='email']", email)
            page.fill("input[name='pass']", password)
            page.click("button[type='submit']")
            logging.info("login_to_facebook: submitted credentials.")

            # Wait for navigation to complete after login
            page.wait_for_timeout(5000)  # Give Facebook time to process login

            if "login" in page.url.lower():
                logging.error("login_to_facebook: still on login page after automated flow.")
                return False
        except Exception as e:
            logging.error(f"login_to_facebook: error during automated login: {e}")
            return False

        # 5) Persist state
        try:
            self.context.storage_state(path=self.facebook_auth_path)
            logging.info("login_to_facebook: session state saved.")
            # Sync to database
            from secret_paths import sync_auth_to_db
            sync_auth_to_db(self.facebook_auth_path, 'facebook')
        except Exception as e:
            logging.warning(f"login_to_facebook: could not save session state: {e}")

        return True
    

    def normalize_facebook_url(self, url: str) -> str:
        """
        If the URL is a Facebook login redirect, unwrap the 'next' parameter and return the real target.
        Otherwise, return the URL unchanged.
        """
        real = canonicalize_facebook_url(url)
        if real != url:
            logging.info(f"normalize_facebook_url: unwrapped login redirect to {real}")
        return real
    

    def navigate_and_maybe_login(self, incoming_url: str, max_attempts: int = 2) -> bool:
        """
        Navigate to a Facebook URL, handle login redirects, detect blocks, and retry.
        Input:
            incoming_url (str): The URL to navigate to.
        Returns:
            bool: True if navigation was successful, False otherwise.
        """
        real_url = self.normalize_facebook_url(incoming_url)
        page = self.logged_in_page

        attempts = max(1, int(max_attempts))
        for attempt in range(attempts):
            try:
                t = random.randint(20000//2, int(20000 * 1.5))
                page.goto(real_url, wait_until="domcontentloaded", timeout=t)
            except PlaywrightTimeoutError:
                logging.warning(f"navigate_and_maybe_login: timeout on {real_url}")

            try:
                state = classify_facebook_access_state(page.url, page.content())
            except Exception:
                state = classify_facebook_access_state(page.url, "")

            if state == "ok":
                return True

            if state == "login":
                logging.info(f"navigate_and_maybe_login: login required for {incoming_url}")
                if not self.login_to_facebook():
                    return False
                # Immediately re-check target URL after login, even in single-attempt mode.
                try:
                    t = random.randint(20000 // 2, int(20000 * 1.5))
                    page.goto(real_url, wait_until="domcontentloaded", timeout=t)
                except PlaywrightTimeoutError:
                    logging.warning(f"navigate_and_maybe_login: timeout after login for {real_url}")
                try:
                    post_login_state = classify_facebook_access_state(page.url, page.content())
                except Exception:
                    post_login_state = classify_facebook_access_state(page.url, "")
                if post_login_state == "ok":
                    return True
                if post_login_state == "blocked":
                    logging.warning(f"navigate_and_maybe_login: blocked or rate-limited on {real_url}")
                    return False
                # Still appears to require login.
                continue

            logging.warning(f"navigate_and_maybe_login: blocked or rate-limited on {real_url}")
            if attempt < attempts - 1:
                cooldown_ms = random.randint(6000, 12000)
                page.wait_for_timeout(cooldown_ms)
                continue
            return False

        return False
    

    def extract_event_links(self, search_url: str) -> set:
        """
        Extracts Facebook event links from a given search URL.
        Handles login if necessary, normalizes the URL, and scrolls to load dynamic content.
        Args:
            search_url (str): The URL to search for events.
        Returns:
            set: A set of unique event links found on the page.
        Logs:
            - Logs the normalized URL.
            - Logs the number of links found.
            - Logs any errors encountered during navigation or extraction.
        """
        # 1) Ensure access
        if not self.navigate_and_maybe_login(search_url):
            logging.error(f"extract_event_links: cannot access {search_url}")
            return set()

        # 2) Normalize for tab logic
        norm_url = self.normalize_facebook_url(search_url)
        logging.info(f"extract_event_links(): Normalized URL: {norm_url}")

        # 3) Handle /events tab
        if "/groups/" in norm_url and not norm_url.rstrip("/").endswith("/events"):
            norm_url = norm_url.rstrip("/") + "/events/"
        self.total_url_attempts += 1
        logging.info(f"extract_event_links(): Navigating to {norm_url}")

        try:
            t = random.randint(20000//2, int(20000 * 1.5))
            self.logged_in_page.goto(norm_url, wait_until="domcontentloaded", timeout=t)
        except Exception as e:
            logging.error(f"extract_event_links(): error loading {norm_url}: {e}")
            return set()

        # scrolling logic unchanged...
        if norm_url.rstrip('/').endswith('/events'):
            self.logged_in_page.wait_for_timeout(2000)
            for _ in range(min(2, self.config['crawling'].get('scroll_depth', 2))):
                self.logged_in_page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                self.logged_in_page.wait_for_timeout(1000)
        else:
            self.logged_in_page.wait_for_timeout(2000)

        html = self.logged_in_page.content()
        links = set(re.findall(r'https://www\.facebook\.com/events/\d+/', html))
        if links:
            self.urls_with_extracted_text += 1
            self.unique_urls.update(links)
        logging.info(f"extract_event_links(): Found {len(links)} links on {norm_url}")

        return links


    def extract_event_text(self, link: str, assume_navigated: bool = False) -> str:
        """
        Extracts the full text content from a Facebook event page.
        Navigates to the specified event link, handling login if necessary, and attempts to load the page content.
        Clicks "See more" buttons to expand hidden text, waits for dynamic content to load, and then extracts all visible text from the page.
        Returns the concatenated text content, or None if the page cannot be accessed or no text is found.
        Args:
            link (str): The URL of the Facebook event page to extract text from.
        Returns:
            str: The extracted text content from the event page, or None if extraction fails.
        """

        if not assume_navigated:
            if not self.navigate_and_maybe_login(link):
                logging.warning(f"extract_event_text: cannot access {link}")
                return None
        
        self.total_url_attempts += 1
        page = self.logged_in_page

        if not page or page.url == "about:blank":
            page = self.context.new_page()
            self.logged_in_page = page
            timeout_value = random.randint(10000, 15000)
            logging.info(f"extract_event_text: Navigating to {link} ({timeout_value} ms)")
            try:
                page.goto(link, wait_until="domcontentloaded", timeout=timeout_value)
            except PlaywrightTimeoutError:
                logging.error(f"extract_event_text: timeout on {link}")
                return None
        
        page.wait_for_timeout(random.randint(4000, 7000))
        for btn in page.query_selector_all("text=/See more/i"):
            try:
                btn.click()
                page.wait_for_timeout(random.randint(3000, 6000))
            except:
                break

        page.wait_for_timeout(random.randint(3000, 5000))
        html = page.content()
        soup = BeautifulSoup(html, 'html.parser')
        full_text = ' '.join(soup.stripped_strings)
        if not full_text:
            logging.warning(f"extract_event_text: no text from {link}")
            return None
        log_extracted_text("extract_event_text", link, full_text, logging.getLogger(__name__))

        return full_text

    
    def extract_relevant_text(self, content: str, link: str) -> str | None:
        """
        Extracts a relevant portion of text from the given content based on specific patterns.

        This function searches for the first occurrence of "More About Discussion" in the content,
        then finds the last occurrence of a day of the week before this phrase, and finally extracts
        the text from this day up to the phrase "Guests See All".

        Args:
            content (str): The text content to be processed.

        Returns:
            str: The extracted relevant text if all patterns are found, otherwise None.

        Logs:
            Logs warnings if any of the required patterns ("More About Discussion", a day of the week,
            or "Guests See All") are not found in the expected positions within the content.
        """
        logging.info(f"def extract_relevant_text(): Extracting relevant text from {link}.")

        # Step 1: Find the first occurrence of "More About Discussion" (case-insensitive)
        mad_pattern = re.compile(r"More About Discussion", re.IGNORECASE)
        mad_match = mad_pattern.search(content)
        if not mad_match:
            logging.warning(f"'More About Discussion' not found in {link}.")
            return None

        mad_start = mad_match.start()

        # Step 2: In the text before "More About Discussion", find the last occurrence of a day of the week
        days_pattern = re.compile(r"\b(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b", re.IGNORECASE)
        days_matches = list(days_pattern.finditer(content, 0, mad_start))
        if not days_matches:
            logging.warning(f"No day of the week found before 'More About Discussion' in {link}.")
            return None

        last_day_match = days_matches[-1]
        day_start = last_day_match.start()

        # Step 3: From the last day match, extract up to "Guests See All"
        gsa_pattern = re.compile(r"Guests See All", re.IGNORECASE)
        gsa_match = gsa_pattern.search(content, last_day_match.end())
        if not gsa_match:
            logging.info(f"def extract_relevant_text(): 'Guests See All' not found after last day of the week in {link}.")
            # Fallback: Extract a reasonable amount of text after the day match to investigate what's actually there
            fallback_end = min(last_day_match.end() + 2000, len(content))  # Extract up to 2000 chars or end of content
            extracted_text = content[day_start:fallback_end]

            # Log the fallback extraction using utility
            logging.info(f"extract_relevant_text: 'Guests See All' not found, using fallback extraction for {link}")
            log_extracted_text("extract_relevant_text", link, extracted_text, logging.getLogger(__name__))

            return extracted_text

        gsa_end = gsa_match.end()

        # Extract the desired text
        extracted_text = content[day_start:gsa_end]

        return extracted_text
    

    def append_df_to_excel(self, df: pd.DataFrame, output_path: str):
        """
        Appends `df` to the first sheet of output_path, creating it if necessary.
        Only called locally, not on Render.
        """
        try:
            # Load existing workbook
            book = load_workbook(output_path)
            with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                # Write df starting at the first empty row (book.active.max_row)
                df.to_excel(
                    writer,
                    index=False,
                    header=False,
                    startrow=book.active.max_row
                )
        except FileNotFoundError:
            # If the file doesn't exist yet, create it with headers
            df.to_excel(output_path, index=False)
    
        
    def scrape_events(self, keywords: list[str], process_callback=None) -> tuple[str, int]:
        """
        Logs into Facebook once, performs searches for keywords, and extracts event link and text.
        Events are processed immediately via callback instead of being accumulated.

        Args:
            keywords (list): List of keywords to search for.
            process_callback (callable): Optional callback function(url, extracted_text, search_url)
                                        called for each event with keywords. If None, events are
                                        accumulated and returned (legacy behavior).

        Returns:
            tuple: The last search_url used and count of events processed.
        """
        self._ensure_authenticated_or_raise()
        base_url = self.config['constants']['fb_base_url']
        location_id = self.config['constants']['fb_location_id']
        events_processed = 0

        for keyword in keywords:
            search_url = f"{base_url}{keyword}{location_id}"
            event_links = self.extract_event_links(search_url)
            logging.info(f"def scrape_events: Used {search_url} to get {len(event_links)} event_links\n")

            self.total_url_attempts += len(event_links)  # Update total URL attempts

            for link in event_links:
                if link in self.urls_visited:
                    continue  # Skip already visited URLs
                else:
                    # Check urls to see if they should be scraped
                    if not db_handler.should_process_url(link):
                        logging.info(f"def scrape_events(): Skipping URL {link} based on historical relevancy.")
                        continue
                    self.unique_urls_count += 1  # Increment unique URL count

                    if len(self.urls_visited) >= self.config['crawling']['urls_run_limit']:
                        logging.info("def scrape_events(): Reached the URL visit limit. Stopping the scraping process.")
                        return search_url, events_processed

                    extracted_text = self.extract_event_text(link)
                    if extracted_text:
                        relevant_text = self.extract_relevant_text(extracted_text, link)
                        if relevant_text:
                            extracted_text = relevant_text
                    self.urls_visited.add(link)

                    if extracted_text:
                        self.urls_with_extracted_text += 1  # Increment URLs with extracted text

                        # Check for keywords in the extracted text
                        found_keywords = [kw for kw in self.keywords_list if kw in extracted_text.lower()]
                        if found_keywords:
                            self.urls_with_found_keywords += 1  # Increment URLs with found keywords
                            logging.info(f"def scrape_events(): Keywords: {found_keywords}: found in text for URL: {link}.")

                            # Process immediately via callback if provided
                            if process_callback:
                                process_callback(link, extracted_text, search_url)

                            events_processed += 1
                            logging.debug(f"Visited URL: {link}. Total visited: {len(self.urls_visited)}")
                        else:
                            logging.info(f"def scrape_events(): No keywords found in extracted text for URL: {link}.")
                    else:
                        logging.info(f"def scrape_events(): No text extracted for URL: {link}.")

        logging.info(f"def scrape_events(): Processed {events_processed} events.")
        return search_url, events_processed


    def process_fb_url(self, url: str, parent_url: str, source: str, keywords: str) -> None:
        """
        Processes a Facebook URL by extracting event information (full or relevant text),
        querying an LLM for event details, and writing results to the database.
        Args:
            url (str): The Facebook URL to process.
            source (str): Source identifier (e.g., 'fb').
            keywords (str): Comma-separated keywords to check for relevance.
        Returns:
            None
        """
        # Canonicalize URL first to avoid persisting login redirect wrappers.
        url = self.normalize_facebook_url(url)

        # Set up url_row for database writing
        url_row = [url, parent_url, source, keywords, False, 1, datetime.now()]

        # Ensure we can access the page
        if not self.navigate_and_maybe_login(url):
            logging.info(f"process_fb_url: cannot access {url}")
            db_handler.write_url_to_db(url_row)
            return

        # Initialize tracking
        self.total_url_attempts += 1

        # 1) Extract text: full event page vs relevant snippet
        if "event" in url:
            extracted_text = self.extract_event_text(url, assume_navigated=True)
        else:
            full_text = self.extract_event_text(url, assume_navigated=True)
            extracted_text = self.extract_relevant_text(full_text, url) if full_text else None

        # 2) Bail if no text
        if not extracted_text:
            logging.info(f"process_fb_url: no text for {url}")
            db_handler.write_url_to_db(url_row)
            return
        self.urls_with_extracted_text += 1

        # 3) Check for keywords
        keywords_found = [kw for kw in self.keywords_list if kw in extracted_text.lower()]
        if not keywords_found:
            logging.info(f"process_fb_url: no keywords in {url}")
            db_handler.write_url_to_db(url_row)
            return
        self.urls_with_found_keywords += 1

        # 4) Query LLM for structured event data
        prompt, schema_type = llm_handler.generate_prompt(url, extracted_text, 'fb')
        if len(prompt) > config['crawling']['prompt_max_length']:
            logging.warning(f"def process_fb_url(): Prompt for URL {url} exceeds maximum length. Skipping LLM query.")
            return
        
        llm_response = llm_handler.query_llm(url, prompt, schema_type)
        if not llm_response or "No events found" in llm_response:
            logging.info(f"process_fb_url: LLM no events for {url}")
            db_handler.write_url_to_db(url_row)
            return

        # 5) Parse JSON and write to DB
        parsed = llm_handler.extract_and_parse_json(llm_response, url, schema_type)
        if not parsed:
            logging.warning(f"process_fb_url: empty LLM response for {url}")
            db_handler.write_url_to_db(url_row)
            return
        events_df = pd.DataFrame(parsed)
        if events_df.empty:
            logging.warning(f"process_fb_url: empty DataFrame for {url}")
            db_handler.write_url_to_db(url_row)
            return

        # Ensure URL column is populated
        if 'url' in events_df.columns and events_df['url'].iloc[0] == '':
            events_df.loc[0, 'url'] = url

        # 6) Write events and mark URL
        db_handler.write_events_to_db(events_df, url, parent_url, source, keywords_found)
        self.events_written_to_db += len(events_df)
    

    def driver_fb_search(self) -> None:
        """Orchestrates the Facebook event search and extraction workflow.
        This method performs the following steps:
            1. Reads keywords from a CSV file specified in the configuration.
            2. Iterates over each set of keywords to construct Facebook search queries.
            3. Scrapes event-related data from the search results.
            4. For each extracted Facebook event URL:
                - Events are processed immediately via streaming callback (no accumulation)
                - Searches for relevant keywords in the extracted text.
                - If keywords are found, processes the event data using an LLM handler and writes results to the database.
            5. Updates the processed status of each keyword set and checkpoints progress to a CSV file.
        Logging is used throughout to track progress, handle crawl limits, and record processing outcomes.
        Returns:
            None
        """
        self._ensure_authenticated_or_raise()
        # Read in keywords and append a column for processed status
        keywords_df = pd.read_csv(self.config['input']['data_keywords'])
        keywords_df['processed'] = keywords_df['processed'] = False

        for idx, row in keywords_df.iterrows():
            keywords_list = row['keywords'].split(',')
            source = row.get('source', '')

            # Define callback to process events immediately (streaming)
            def process_event(url: str, extracted_text: str, parent_url: str):
                """Callback to process each event immediately as it's extracted."""
                logging.info(f"def driver_fb_search(): Processing Facebook URL: {url}")

                if len(self.urls_visited) >= self.config['crawling']['urls_run_limit']:
                    logging.info("def driver_fb_search(): Reached crawl limit. Stopping processing.")
                    return

                # Check for keywords in the extracted text
                found_keywords = [kw for kw in self.keywords_list if kw in extracted_text.lower()]
                found_keywords = ', '.join(found_keywords)
                if found_keywords:
                    logging.info(f"def driver_fb_search(): Keywords found in text for {url}.")

                    # Set prompt and process LLM response immediately
                    prompt_type = 'fb'
                    llm_response = llm_handler.process_llm_response(url, parent_url, extracted_text, source, found_keywords, prompt_type)

                    # If events were successfully extracted and written to the DB
                    if llm_response:
                        self.events_written_to_db += 1
                        logging.info(f"def driver_fb_search(): Events successfully written to DB for {url}.")
                    else:
                        logging.warning(f"def driver_fb_search(): No events extracted for {url}.")
                        url_row = [url, parent_url, source, found_keywords, False, 1, datetime.now()]
                        db_handler.write_url_to_db(url_row)
                else:
                    keywords = ''
                    url_row = [url, parent_url, source, keywords, False, 1, datetime.now()]
                    db_handler.write_url_to_db(url_row)
                    logging.info(f"def driver_fb_search(): No keywords found in extracted text for URL: {url}.")

            # Scrape events with streaming callback (no accumulation)
            search_url, events_processed = self.scrape_events(keywords_list, process_callback=process_event)
            logging.info(f"def driver_fb_search(): Processed {events_processed} events from search_url: {search_url}.")

            # Checkpoint the keywords (only locally, not on Render)
            keywords_df.loc[idx, 'processed'] = True
            if os.getenv('RENDER') != 'true':
                keywords_df.to_csv(self.config['checkpoint']['fb_search'], index=False)
                logging.info(f"def driver_fb_search(): Keywords checkpoint updated.")
            else:
                logging.info(f"def driver_fb_search(): Skipping checkpoint write on Render")


    def driver_fb_urls(self) -> None:
        """
        1. Gets all of the URLs from the urls table where the link is like '%facebook%'.
        2. For each URL, processes it and then scrapes any event links by hitting the /events/ subpage.
        3. Writes every processed URL—including event links—to the checkpoint CSV.
        """
        self._ensure_authenticated_or_raise()
        # 1) Load or initialize the checkpoint dataframe
        # On Render, always query database fresh (no checkpointing)
        is_render = os.getenv('RENDER') == 'true'
        query = text("""
            SELECT *
            FROM urls
            WHERE link ILIKE :link_pattern
              AND link NOT ILIKE :exclude_sharer
              AND link NOT ILIKE :exclude_dialog
              AND link NOT ILIKE :exclude_recover
              AND link NOT ILIKE :exclude_plugins
              AND link NOT ILIKE :exclude_share_php
        """)
        params = {
            'link_pattern': '%facebook%',
            'exclude_sharer': '%/sharer/%',
            'exclude_dialog': '%/dialog/%',
            'exclude_recover': '%/recover/%',
            'exclude_plugins': '%/plugins/%',
            'exclude_share_php': '%share.php%',
        }
        if is_render:
            fb_urls_df = pd.read_sql(query, db_handler.conn, params=params)
            logging.info(f"def driver_fb_urls(): Retrieved {fb_urls_df.shape[0]} Facebook URLs from the database.")
        elif should_use_fb_checkpoint(config, is_render):
            fb_urls_df = pd.read_csv(config['checkpoint']['fb_urls_cp'])
            logging.info(f"def driver_fb_urls(): Loaded {fb_urls_df.shape[0]} Facebook URLs from checkpoint CSV.")
        else:
            fb_urls_df = pd.read_sql(query, db_handler.conn, params=params)
            logging.info(f"def driver_fb_urls(): Retrieved {fb_urls_df.shape[0]} Facebook URLs from the database.")

        fb_urls_df, seed_stats = sanitize_facebook_seed_urls(fb_urls_df)
        logging.info(
            "def driver_fb_urls(): Seed URL sanitize stats: input=%s output=%s canonicalized=%s dropped_non_content=%s dropped_empty=%s dropped_duplicates=%s",
            seed_stats["input_rows"],
            seed_stats["output_rows"],
            seed_stats["canonicalized_rows"],
            seed_stats["non_content_rows_dropped"],
            seed_stats["empty_rows_dropped"],
            seed_stats["duplicate_rows_dropped"],
        )

        # 2) Add checkpoint columns
        fb_urls_df['processed'] = False
        fb_urls_df['events_processed'] = False
        if os.getenv('RENDER') != 'true':
            fb_urls_df.to_csv(self.config['checkpoint']['fb_urls'], index=False)
        checkpoint_write_every = int(self.config.get('checkpoint', {}).get('fb_urls_write_every', 10) or 10)
        checkpoint_updates = 0
        checkpoint_dirty = False
        existing_links = set(fb_urls_df['link'].dropna().astype(str))

        def flush_checkpoint(force: bool = False) -> None:
            nonlocal checkpoint_updates, checkpoint_dirty
            if is_render or not checkpoint_dirty:
                return
            if force or checkpoint_updates >= checkpoint_write_every:
                fb_urls_df.to_csv(self.config['checkpoint']['fb_urls'], index=False)
                checkpoint_updates = 0
                checkpoint_dirty = False

        def mark_checkpoint_dirty() -> None:
            nonlocal checkpoint_updates, checkpoint_dirty
            if is_render:
                return
            checkpoint_dirty = True
            checkpoint_updates += 1
            flush_checkpoint(force=False)

        # 3) Iterate each base Facebook URL
        if fb_urls_df.shape[0] > 0:
            processed_base_urls = set()
            for idx, row in fb_urls_df.iterrows():
                raw_base_url = row['link']
                base_url = self.normalize_facebook_url(raw_base_url)
                parent_url = row['parent_url']
                source = row['source']
                keywords = row['keywords']
                logging.info(f"def driver_fb_urls(): Processing base URL: {base_url}")

                if is_facebook_login_redirect(raw_base_url):
                    logging.info(
                        "def driver_fb_urls(): Canonicalized login redirect base URL %s -> %s",
                        raw_base_url,
                        base_url,
                    )

                if is_non_content_facebook_url(base_url):
                    logging.info(
                        "def driver_fb_urls(): Skipping non-content Facebook endpoint: %s",
                        base_url,
                    )
                    url_row = [base_url, parent_url, source, keywords, False, 1, datetime.now()]
                    db_handler.write_url_to_db(url_row)
                    fb_urls_df.loc[fb_urls_df['link'] == raw_base_url, 'processed'] = True
                    fb_urls_df.loc[fb_urls_df['link'] == raw_base_url, 'events_processed'] = True
                    mark_checkpoint_dirty()
                    continue

                # Skip if already done
                if base_url in self.urls_visited or base_url in processed_base_urls:
                    continue

                # Check urls to see if they should be scraped
                if not db_handler.should_process_url(base_url):
                    logging.info(f"def eventbrite_search(): Skipping URL {base_url} based on historical relevancy.")
                    continue

                # Track whether this base URL yields events either directly or via /events/.
                base_events_before = self.events_written_to_db
                base_yielded_events = False

                # Process the base URL itself (writes any events found on that exact page)
                self.process_fb_url(base_url, parent_url, source, keywords)
                if self.events_written_to_db > base_events_before:
                    base_yielded_events = True
                self.urls_visited.add(base_url)
                processed_base_urls.add(base_url)

                # Mark as processed
                fb_urls_df.loc[fb_urls_df['link'] == raw_base_url, 'processed'] = True
                fb_urls_df.loc[fb_urls_df['link'] == base_url, 'processed'] = True
                mark_checkpoint_dirty()
                if os.getenv('RENDER') != 'true':
                    logging.info(f"def driver_fb_urls(): Base URL marked processed: {base_url}")
                else:
                    logging.info(f"def driver_fb_urls(): Skipping checkpoint write on Render")

                # Honor the run limit
                if len(self.urls_visited) >= self.config['crawling']['urls_run_limit']:
                    break

                # 4) Now scrape *all* event links by auto-navigating to /events/
                fb_event_links = self.extract_event_links(base_url)
                if not fb_event_links:
                    logging.info(f"driver_fb_urls(): No events tab or no events found on {base_url}")

                # 5) Process each event link
                for event_url in fb_event_links:
                    event_url = self.normalize_facebook_url(event_url)

                    if is_non_content_facebook_url(event_url):
                        logging.info(
                            "def driver_fb_urls(): Skipping non-content discovered Facebook endpoint: %s",
                            event_url,
                        )
                        url_row = [event_url, base_url, source, keywords, False, 1, datetime.now()]
                        db_handler.write_url_to_db(url_row)
                        continue

                    if event_url in self.urls_visited:
                        continue
                    
                    # Check urls to see if they should be scraped
                    if not db_handler.should_process_url(event_url):
                        logging.info(f"def eventbrite_search(): Skipping URL {event_url} based on historical relevancy.")
                        continue

                    event_events_before = self.events_written_to_db
                    self.process_fb_url(event_url, base_url, source, keywords)
                    if self.events_written_to_db > event_events_before:
                        base_yielded_events = True
                    self.urls_visited.add(event_url)

                    # Add or update checkpoint row
                    if event_url not in existing_links:
                        new_row = pd.DataFrame({
                            'link': [event_url],
                            'source': [source],
                            'keywords': [keywords],
                            'processed': [True],
                            'events_processed': [True]
                        })
                        fb_urls_df = pd.concat([fb_urls_df, new_row], ignore_index=True)
                        existing_links.add(event_url)
                    else:
                        fb_urls_df.loc[fb_urls_df['link'] == event_url, 'processed'] = True
                        fb_urls_df.loc[fb_urls_df['link'] == event_url, 'events_processed'] = True

                    mark_checkpoint_dirty()
                    if os.getenv('RENDER') != 'true':
                        logging.info(f"def driver_fb_urls(): Event URL marked processed: {event_url}")
                    else:
                        logging.info(f"def driver_fb_urls(): Skipping checkpoint write on Render")

                    if len(self.urls_visited) >= self.config['crawling']['urls_run_limit']:
                        break

                # 6) Finally mark that we've scraped events for the base URL
                fb_urls_df.loc[fb_urls_df['link'] == raw_base_url, 'events_processed'] = True
                fb_urls_df.loc[fb_urls_df['link'] == base_url, 'events_processed'] = True
                mark_checkpoint_dirty()
                if os.getenv('RENDER') != 'true':
                    logging.info(f"def driver_fb_urls(): Events_scraped flag set for base URL: {base_url}")
                else:
                    logging.info(f"def driver_fb_urls(): Skipping checkpoint write on Render")

                # If any downstream event was yielded, persist an explicit relevant=True
                # row for the base URL so validation reflects actual extraction success.
                if base_yielded_events:
                    db_handler.write_url_to_db([base_url, parent_url, source, keywords, True, 1, datetime.now()])
                    logging.info(
                        "def driver_fb_urls(): Base URL marked relevant due to yielded events: %s",
                        base_url,
                    )

        else:
            logging.warning("def driver_fb_urls(): No Facebook URLs returned from the database.")
        flush_checkpoint(force=True)


    def write_run_statistics(self) -> None:
        """
        Writes run statistics to the database.
        """
        end_time = datetime.now()
        logging.info(f"write_run_statistics(): Writing run statistics for {self.run_name}.")

        elapsed_time = str(end_time - self.start_time)
        time_stamp = datetime.now()

        run_data = pd.DataFrame([{
            "run_name": self.run_name,
            "run_description": self.run_description,
            "start_time": self.start_time,
            "end_time": end_time,
            "elapsed_time": elapsed_time,
            "python_file_name": "fb.py",
            "unique_urls_count": len(self.urls_visited),
            "total_url_attempts": self.total_url_attempts,
            "urls_with_extracted_text": self.urls_with_extracted_text,
            "urls_with_found_keywords": self.urls_with_found_keywords,
            "events_written_to_db": self.events_written_to_db,
            "time_stamp": time_stamp
        }])

        run_data.to_sql("runs", db_handler.get_db_connection(), if_exists="append", index=False)
        logging.info(f"write_run_statistics(): Run statistics written to database for {self.run_name}.")


    def checkpoint_events(self) -> None:
        """
        Checkpoints the events by reading from the checkpoint CSV file and writing to the database.
        """
        logging.info("def checkpoint_events(): Starting to checkpoint events.")

        # Read in checkpoint file
        df = pd.read_csv('checkpoint/extracted_text_may_29_2025.csv')
        if df.empty:
            logging.warning("checkpoint_events(): No data found in the checkpoint file.")
            return

        parent_url = ''
        source = 'checkpoint'
        relevant = False
        crawl_try = 1

        for _, row in df.iterrows():
            url = row['url']
            extracted_text = row['extracted_text']

            # try to narrow down to just the relevant snippet
            relevant_text = fb_scraper.extract_relevant_text(extracted_text, url)
            if not relevant_text:
                relevant_text = extracted_text

            # skip anything with no keywords
            found_keywords = [kw for kw in self.keywords_list if kw in extracted_text.lower()]
            if not found_keywords:
                logging.info(f"checkpoint_events(): No keywords found in extracted text for URL: {url}.")
                continue

            found_keywords = ', '.join(found_keywords)
            logging.info(f"checkpoint_events(): Keywords found in text for {url}: {found_keywords}.")

            # prepare the URL‐row metadata (note the timestamp call)
            url_row = [
                url,
                parent_url,
                source,
                found_keywords,
                relevant,
                crawl_try,
                datetime.now()
            ]

            # 4) Query the LLM
            prompt, schema_type = llm_handler.generate_prompt(url, relevant_text, 'fb')
            llm_response = llm_handler.query_llm(url, prompt, schema_type)
            if not llm_response or "No events found" in llm_response:
                logging.info(f"checkpoint_events(): LLM returned no events for {url}")
                db_handler.write_url_to_db(url_row)
                continue

            # 5) Parse JSON, build DataFrame
            parsed = llm_handler.extract_and_parse_json(llm_response, url, schema_type)
            if not parsed:
                logging.warning(f"checkpoint_events(): empty LLM response for {url}")
                db_handler.write_url_to_db(url_row)
                continue

            events_df = pd.DataFrame(parsed)
            if events_df.empty:
                logging.warning(f"checkpoint_events(): empty DataFrame for {url}")
                db_handler.write_url_to_db(url_row)
                continue

            # ensure the URL column is filled
            if 'url' in events_df.columns and events_df['url'].iloc[0] == '':
                events_df.loc[0, 'url'] = url

            # 6) write events and tally
            db_handler.write_events_to_db(events_df, url, parent_url, source, found_keywords)
            self.events_written_to_db += len(events_df)

        logging.info(f"checkpoint_events(): Completed. Wrote {self.events_written_to_db} events.")


def main():
    """ Main function to initialize and run the Facebook scraper. """
    with open('config/config.yaml', 'r') as file:
        config = yaml.safe_load(file)

    # Setup centralized logging
    from logging_config import setup_logging
    setup_logging('fb')
    logging.info("\n\nfb.py starting...")
    logging.info("__main__: fb.py revision %s", get_git_revision())

    start_time = datetime.now()
    logging.info(f"\n\n__main__: Starting the crawler process at {start_time}")

    # Initialize handlers - make them global for class methods
    global llm_handler, db_handler
    llm_handler = LLMHandler(config_path='config/config.yaml')
    db_handler = llm_handler.db_handler  # Use the DatabaseHandler from LLMHandler

    # Get the file name of the running script
    file_name = os.path.basename(__file__)

    # Count events and URLs before running
    start_df = db_handler.count_events_urls_start(file_name)

    # Initialize scraper
    fb_scraper = FacebookEventScraper(config_path='config/config.yaml')

    # Run
    fb_scraper.driver_fb_search()
    fb_scraper.driver_fb_urls()

    # To be removed in production
    # fb_scraper.checkpoint_events()

    # Write run statistics to the database
    fb_scraper.write_run_statistics()

    # Close browser and Playwright
    fb_scraper.browser.close()
    fb_scraper.playwright.stop()

    # Count events and URLs after running
    db_handler.count_events_urls_end(start_df, file_name)

    # End time and elapsed time logging
    end_time = datetime.now()
    logging.info(f"__main__: Finished the crawler process at {end_time}")
    logging.info(f"__main__: Total time taken: {end_time - start_time}")


if __name__ == "__main__":
    main()
